import sqlite3
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
db_path = BASE_DIR / "database" / "lease_renewal.db"
output_path = BASE_DIR / "excel" / "Tenant_Renewal_Management_Report_Upgraded.xlsx"

if not db_path.exists():
    raise FileNotFoundError(f"SQLite database not found: {db_path}")

query = """
SELECT
    tenant_id,
    property,
    unit_no,
    tenant_name,
    basic_rent,
    current_lease_start_date,
    current_lease_end_date,
    lease_type,
    renewal_status,
    payment_score_overall,
    payment_score_basic_rental,
    payment_score_utilities,
    arrears_incidents_calc,
    lease_longevity_months,
    deposit_coverage_score,
    tenant_risk_score,
    tenant_grade,
    renewal_recommendation
FROM tenant_risk_scorecard
ORDER BY tenant_risk_score DESC;
"""

conn = sqlite3.connect(db_path)
df = pd.read_sql_query(query, conn)
conn.close()

# Management category datasets.
strong = df[df["renewal_recommendation"].isin(["Strongly Renew", "Renew"])].copy()
review = df[df["renewal_recommendation"].isin(["Renew with Review"])].copy()
management = df[df["renewal_recommendation"].isin(["Management Review"])].copy()
do_not = df[df["renewal_recommendation"].isin(["Do Not Renew"])].copy()

summary = df.groupby(["tenant_grade", "renewal_recommendation"], dropna=False).size().reset_index(name="tenant_count")
priority = df.sort_values(["tenant_risk_score", "arrears_incidents_calc"], ascending=[True, False]).head(15).copy()

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="All Tenants", index=False)
    strong.to_excel(writer, sheet_name="Strong Renewals", index=False)
    review.to_excel(writer, sheet_name="Renew with Review", index=False)
    management.to_excel(writer, sheet_name="Management Review", index=False)
    do_not.to_excel(writer, sheet_name="Do Not Renew", index=False)
    summary.to_excel(writer, sheet_name="Grade Summary", index=False)
    priority.to_excel(writer, sheet_name="Priority Action List", index=False)

wb = load_workbook(output_path)

header_fill = PatternFill("solid", fgColor="1F4E78")
subheader_fill = PatternFill("solid", fgColor="D9EAF7")
white_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E2F3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    max_col = ws.max_column
    max_row = ws.max_row

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        values = [str(ws.cell(row=row, column=col_idx).value or "") for row in range(1, min(max_row, 50) + 1)]
        width = min(max(max(len(v) for v in values) + 2, 12), 32)
        ws.column_dimensions[col_letter].width = width

    if "tenant_risk_score" in [cell.value for cell in ws[1]]:
        score_col = [cell.value for cell in ws[1]].index("tenant_risk_score") + 1
        score_letter = get_column_letter(score_col)
        ws.conditional_formatting.add(
            f"{score_letter}2:{score_letter}{max_row}",
            ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
        )

    if "tenant_grade" in [cell.value for cell in ws[1]]:
        grade_col = [cell.value for cell in ws[1]].index("tenant_grade") + 1
        grade_letter = get_column_letter(grade_col)
        for grade_value, fill_color in [("A", "63BE7B"), ("B", "A9D18E"), ("C", "FFEB84"), ("D", "F4B183"), ("F", "F8696B")]:
            ws.conditional_formatting.add(
                f"{grade_letter}2:{grade_letter}{max_row}",
                CellIsRule(operator="equal", formula=[f'"{grade_value}"'], fill=PatternFill("solid", fgColor=fill_color)),
            )

# Dashboard sheet.
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False

dash["A1"] = "Tenant Renewal Risk Scorecard"
dash["A1"].font = Font(size=18, bold=True, color="1F4E78")
dash.merge_cells("A1:H1")

dash["A3"] = "Total Tenants"
dash["B3"] = len(df)
dash["A4"] = "Average Risk Score"
dash["B4"] = round(float(df["tenant_risk_score"].mean()), 1) if len(df) else 0
dash["A5"] = "Strong Renew / Renew"
dash["B5"] = len(strong)
dash["A6"] = "Management Review / Do Not Renew"
dash["B6"] = len(management) + len(do_not)

for cell in dash["A3:A6"]:
    cell[0].fill = subheader_fill
    cell[0].font = Font(bold=True)
for cell in dash["B3:B6"]:
    cell[0].font = Font(bold=True)
    cell[0].alignment = Alignment(horizontal="center")

# Grade summary table on dashboard.
dash["A9"] = "Grade"
dash["B9"] = "Tenant Count"
for cell in dash[9]:
    cell.fill = header_fill
    cell.font = white_font
    cell.border = border

grade_counts = df["tenant_grade"].value_counts().reindex(["A", "B", "C", "D", "F"]).fillna(0).astype(int)
for idx, (grade, count) in enumerate(grade_counts.items(), start=10):
    dash.cell(row=idx, column=1).value = grade
    dash.cell(row=idx, column=2).value = int(count)
    dash.cell(row=idx, column=1).border = border
    dash.cell(row=idx, column=2).border = border

# Recommendation summary table.
dash["D9"] = "Recommendation"
dash["E9"] = "Tenant Count"
dash["D9"].fill = header_fill
dash["D9"].font = white_font
dash["E9"].fill = header_fill
dash["E9"].font = white_font
rec_counts = df["renewal_recommendation"].value_counts()
for idx, (rec, count) in enumerate(rec_counts.items(), start=10):
    dash.cell(row=idx, column=4).value = rec
    dash.cell(row=idx, column=5).value = int(count)
    dash.cell(row=idx, column=4).border = border
    dash.cell(row=idx, column=5).border = border

# Charts.
bar = BarChart()
bar.title = "Tenant Count by Grade"
bar.y_axis.title = "Tenant Count"
bar.x_axis.title = "Grade"
data = Reference(dash, min_col=2, min_row=9, max_row=14)
cats = Reference(dash, min_col=1, min_row=10, max_row=14)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.height = 7
bar.width = 12
dash.add_chart(bar, "A17")

pie = PieChart()
pie.title = "Renewal Recommendation Split"
max_rec_row = 9 + len(rec_counts)
pie_data = Reference(dash, min_col=5, min_row=9, max_row=max_rec_row)
pie_cats = Reference(dash, min_col=4, min_row=10, max_row=max_rec_row)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie.height = 7
pie.width = 12
dash.add_chart(pie, "D17")

for col in range(1, 9):
    dash.column_dimensions[get_column_letter(col)].width = 22

wb.save(output_path)

print("Upgraded management report exported successfully.")
print(f"File created at: {output_path}")
